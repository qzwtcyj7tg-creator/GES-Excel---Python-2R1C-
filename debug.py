from __future__ import annotations

from pathlib import Path
from typing import Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


class ValidierungsExcelErsteller:
    def __init__(self,
                 referenz_excel_pfad: str | Path,
                 export_ordner: str | Path | None = None,
                 output_ordner: Optional[str | Path] = None,
                 suchordner: str | Path | None = None):
        self.referenz_excel_pfad = Path(referenz_excel_pfad)
        ordner = suchordner if suchordner is not None else export_ordner
        if ordner is None:
            raise ValueError('Bitte export_ordner oder suchordner angeben.')

        self.suchordner = Path(ordner)
        self.output_ordner = Path(output_ordner) if output_ordner else self.suchordner
        self.output_ordner.mkdir(parents=True, exist_ok=True)

        self._pairs = [
            ('Nutzungssignal (Simulation) [-]', 'Nutzungssignal Excel', 'd Nutzungssignal'),
            ('Zulufttemperatur [°C]', 'Zuluft Excel', 'dZuluft'),
            ('Volumenstrom [m3/h]', 'Volumenstrom Excel', 'Differenz Volumenstrom'),
            ('Heizregister_Leistung [W]', 'NHR Leistung', 'dNHR'),
            ('Heizung [W]', 'Heizung Excel', 'dHeizung'),
            ('Innentemp [°C]', 'Innentemperatur Excel', 'dInnentemp'),
            ('Solare Einstrahlung [W]', 'Solare Einstrahlung Excel', 'd Solare Eintrahlung'),
            ('Temp nach WRG [°C]', 'Temp nach WRG', 'dTempWRG'),
            ('Temp Abl [°C]', 'Temp Abl', 'dAbl'),
            ('HV', 'HV Excel', 'dHV'),
            ('Theta Test', 'Theta Test Excel', 'dTheta Test Excel'),
            ('Theta 0W', 'Theta 0W Excel', 'dTheta 0W'),
            ('T Soll', 'T Soll Excel', 'd Tsoll'),
        ]

        self._ziel_spalten = [
            'Stunde',
            'Außentemperatur [°C]',
            'Nutzungssignal (Simulation) [-]', 'Nutzungssignal Excel', 'd Nutzungssignal',
            'Zulufttemperatur [°C]', 'Zuluft Excel', 'dZuluft',
            'Volumenstrom [m3/h]', 'Volumenstrom Excel', 'Differenz Volumenstrom',
            'Heizregister_Leistung [W]', 'NHR Leistung', 'dNHR',
            'Heizung [W]', 'Heizung Excel', 'dHeizung',
            'Innentemp [°C]', 'Innentemperatur Excel', 'dInnentemp',
            'Solare Einstrahlung [W]', 'Solare Einstrahlung Excel', 'd Solare Eintrahlung',
            'Temp nach WRG [°C]', 'Temp nach WRG', 'dTempWRG',
            'Temp Abl [°C]', 'Temp Abl', 'dAbl',
            'HV', 'HV Excel', 'dHV',
            'Theta Test', 'Theta Test Excel', 'dTheta Test Excel',
            'Theta 0W', 'Theta 0W Excel', 'dTheta 0W',
            'T Soll', 'T Soll Excel', 'd Tsoll'
        ]

    def _neueste_exportdatei_finden(self) -> Path:
        if not self.suchordner.exists():
            raise FileNotFoundError(f'Suchordner existiert nicht: {self.suchordner}')

        muster = [
            'Zuluft_Ergebnisse_*.xlsx', 'Zuluft_Ergebnisse_*.xls', 'Zuluft_Ergebnisse_*.excel',
            '*Zuluft*Ergebnisse*.xlsx', '*Zuluft*Ergebnisse*.xls', '*Zuluft*Ergebnisse*.excel'
        ]
        kandidaten = []
        for m in muster:
            kandidaten.extend(self.suchordner.glob(m))

        kandidaten = [
            p for p in kandidaten
            if p.is_file()
            and not p.name.startswith('~$')
            and p.resolve() != self.referenz_excel_pfad.resolve()
        ]
        kandidaten = sorted(set(kandidaten), key=lambda p: p.stat().st_mtime, reverse=True)

        if not kandidaten:
            vorhandene = [p.name for p in self.suchordner.iterdir() if p.is_file()]
            raise FileNotFoundError(
                f'Keine passende Exportdatei in {self.suchordner} gefunden.\n'
                'Erwartet wurde z. B. Zuluft_Ergebnisse_20260411.xlsx.\n'
                f'Gefundene Dateien: {vorhandene[:30]}'
            )
        return kandidaten[0]

    @staticmethod
    def _normiere(df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        df.columns = [str(c).strip() for c in df.columns]
        return df

    def _daten_laden(self):
        py_datei = self._neueste_exportdatei_finden()
        df_py = pd.read_excel(py_datei)
        df_ref = pd.read_excel(self.referenz_excel_pfad)
        return self._normiere(df_py), self._normiere(df_ref), py_datei

    def _aufbauen(self, df_py: pd.DataFrame, df_ref: pd.DataFrame) -> pd.DataFrame:
        n = min(len(df_py), len(df_ref))
        df_py = df_py.iloc[:n].reset_index(drop=True)
        df_ref = df_ref.iloc[:n].reset_index(drop=True)
        result = pd.DataFrame(index=range(n))

        result['Stunde'] = df_py['Stunde'] if 'Stunde' in df_py.columns else df_ref['Stunde']
        result['Außentemperatur [°C]'] = df_py['Außentemperatur [°C]'] if 'Außentemperatur [°C]' in df_py.columns else df_ref['Außentemperatur [°C]']

        for py_col, ref_col, diff_col in self._pairs:
            py_vals = df_py[py_col] if py_col in df_py.columns else pd.Series([pd.NA] * n)
            ref_vals = df_ref[ref_col] if ref_col in df_ref.columns else pd.Series([pd.NA] * n)
            result[py_col] = py_vals
            result[ref_col] = ref_vals
            result[diff_col] = pd.to_numeric(py_vals, errors='coerce') - pd.to_numeric(ref_vals, errors='coerce')

        for c in self._ziel_spalten:
            if c not in result.columns:
                result[c] = pd.NA
        return result[self._ziel_spalten]

    def _formatieren(self, datei: Path):
        wb = load_workbook(datei)
        ws = wb.active

        # Farben definieren
        green_fill = PatternFill(fill_type='solid', start_color='90EE90', end_color='90EE90')  # Hellgrün für 0 + Rundungsfehler
        light_orange_fill = PatternFill(fill_type='solid', start_color='FFF2CC', end_color='FFF2CC')  # Hellgelb für 0-0.3
        orange_fill = PatternFill(fill_type='solid', start_color='F4B183', end_color='F4B183')  # Orange für 0.3-1
        red_fill = PatternFill(fill_type='solid', start_color='FF6666', end_color='FF6666')  # Rot für >1

        header = [cell.value for cell in ws[1]]
        diff_spalten = []
        for i, name in enumerate(header, start=1):
            if isinstance(name, str):
                n = name.strip()
                if n.startswith('d') or n == 'Differenz Volumenstrom':
                    diff_spalten.append(i)

        for row in range(2, ws.max_row + 1):
            for col_idx in diff_spalten:
                cell = ws.cell(row=row, column=col_idx)
                try:
                    val = float(cell.value)
                    abs_val = abs(val)
                except (TypeError, ValueError):
                    continue

                # Rundungsfehler (sehr kleine Werte < 1e-10) als 0 behandeln
                if abs_val < 1e-10:  # ← Rundungsfehler-Grenze
                    cell.fill = green_fill
                elif abs_val == 0:
                    cell.fill = green_fill
                elif abs_val <= 0.3:
                    cell.fill = light_orange_fill
                elif abs_val <= 1:
                    cell.fill = orange_fill
                else:  # > 1
                    cell.fill = red_fill

        # Spaltenbreite automatisch an Inhalt anpassen
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max 50 Zeichen Breite
            ws.column_dimensions[column_letter].width = adjusted_width

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        wb.save(datei)

    def excel_erstellen(self, dateiname: Optional[str] = None) -> Path:
        df_py, df_ref, py_datei = self._daten_laden()
        result = self._aufbauen(df_py, df_ref)

        if not dateiname:
            stem = py_datei.stem
            if stem.startswith('Validierung_'):
                dateiname = f'{stem}.xlsx'
            else:
                dateiname = f'Validierung_{stem}.xlsx'

        ziel = self.output_ordner / dateiname
        result.to_excel(ziel, index=False)
        self._formatieren(ziel)
        return ziel


if __name__ == '__main__':
    basisordner = Path(r"C:\Users\nicol\OneDrive - Technische Hochschule Augsburg\THA\M12\2025\Python_Modell_2R1C\export")
    referenz = basisordner / 'Ergebnise_Vergleich_excel_Stand_110426.xlsx'
    validierungsordner = basisordner / 'Validierung'

    ersteller = ValidierungsExcelErsteller(
        referenz_excel_pfad=referenz,
        suchordner=basisordner,
        output_ordner=validierungsordner,
    )
    datei = ersteller.excel_erstellen()
    print(datei)
